import openpyxl
import os

def crear_archivo_si_no_existe(nombre_archivo, encabezados):
    if not os.path.exists(nombre_archivo):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(encabezados)
        wb.save(nombre_archivo)

def guardar_en_excel(datos, nombre_archivo):
    crear_archivo_si_no_existe(nombre_archivo, encabezados=[])
    wb = openpyxl.load_workbook(nombre_archivo)
    ws = wb.active
    ws.append(datos)
    wb.save(nombre_archivo)

def leer_excel(nombre_archivo):
    if not os.path.exists(nombre_archivo):
        return []

    wb = openpyxl.load_workbook(nombre_archivo)
    ws = wb.active
    return [list(row) for row in ws.iter_rows(values_only=True)][1:]